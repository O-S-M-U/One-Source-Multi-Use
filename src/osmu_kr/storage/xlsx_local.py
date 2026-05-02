"""LocalXlsxStorage — Excel/Numbers에서 직접 열 수 있는 .xlsx 백엔드."""
from __future__ import annotations

import os
import threading
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from ..models import ContentRecord, KeywordPoolItem, ResearchHistoryRecord
from .base import BaseStorage


class LocalXlsxStorage(BaseStorage):
    name = "xlsx"
    POOL_SHEET = "keyword_pool"
    CONTENT_SHEET = "content_db"
    HISTORY_SHEET = "research_history"

    def __init__(self, data_dir: str = "./data", filename: str = "osmu_workbook.xlsx"):
        self.data_dir = data_dir
        os.makedirs(self.data_dir, exist_ok=True)
        self.path = os.path.join(self.data_dir, filename)
        self._lock = threading.Lock()
        self._ensure_workbook()

    def _ensure_workbook(self) -> None:
        if not os.path.isfile(self.path) or os.path.getsize(self.path) == 0:
            wb = Workbook()
            wb.remove(wb.active)
            ws_p = wb.create_sheet(self.POOL_SHEET); ws_p.append(KeywordPoolItem.HEADER)
            ws_c = wb.create_sheet(self.CONTENT_SHEET); ws_c.append(ContentRecord.HEADER)
            ws_h = wb.create_sheet(self.HISTORY_SHEET); ws_h.append(ResearchHistoryRecord.HEADER)
            self._auto_fit(ws_p, KeywordPoolItem.HEADER)
            self._auto_fit(ws_c, ContentRecord.HEADER)
            self._auto_fit(ws_h, ResearchHistoryRecord.HEADER)
            wb.save(self.path)
            return
        wb = load_workbook(self.path)
        changed = False
        if self.POOL_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(self.POOL_SHEET); ws.append(KeywordPoolItem.HEADER)
            self._auto_fit(ws, KeywordPoolItem.HEADER); changed = True
        if self.CONTENT_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(self.CONTENT_SHEET); ws.append(ContentRecord.HEADER)
            self._auto_fit(ws, ContentRecord.HEADER); changed = True
        if self.HISTORY_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(self.HISTORY_SHEET); ws.append(ResearchHistoryRecord.HEADER)
            self._auto_fit(ws, ResearchHistoryRecord.HEADER); changed = True
        if changed:
            wb.save(self.path)

    @staticmethod
    def _auto_fit(ws, header: list) -> None:
        for i, name in enumerate(header, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(name)) + 4)

    def _load(self):
        return load_workbook(self.path)

    def _read_rows(self, sheet_name: str) -> List[list]:
        wb = self._load()
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None or c == "" for c in row):
                continue
            rows.append(["" if c is None else c for c in row])
        return rows

    def _write_all(self, sheet_name: str, header: list, rows: List[list]) -> None:
        with self._lock:
            wb = self._load()
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            ws.append(header)
            for r in rows:
                ws.append(r)
            self._auto_fit(ws, header)
            order = [self.POOL_SHEET, self.CONTENT_SHEET]
            for i, name in enumerate(order):
                if name in wb.sheetnames:
                    wb.move_sheet(name, offset=i - wb.sheetnames.index(name))
            wb.save(self.path)

    def _append_row(self, sheet_name: str, row: list) -> None:
        with self._lock:
            wb = self._load()
            ws = wb[sheet_name]
            ws.append(row)
            wb.save(self.path)

    def list_pool(self):
        return [KeywordPoolItem.from_row(r) for r in self._read_rows(self.POOL_SHEET)]

    def get_pool(self, keyword_id):
        for it in self.list_pool():
            if it.keyword_id == keyword_id:
                return it
        return None

    def upsert_pool(self, item):
        items = self.list_pool()
        for i, it in enumerate(items):
            if it.keyword_id == item.keyword_id:
                items[i] = item
                break
        else:
            items.append(item)
        self.replace_pool(items)

    def delete_pool(self, keyword_id):
        items = self.list_pool()
        new_items = [it for it in items if it.keyword_id != keyword_id]
        if len(new_items) == len(items):
            return False
        self.replace_pool(new_items)
        return True

    def replace_pool(self, items):
        self._write_all(self.POOL_SHEET, KeywordPoolItem.HEADER, [it.to_row() for it in items])

    def list_content(self):
        return [ContentRecord.from_row(r) for r in self._read_rows(self.CONTENT_SHEET)]

    def append_content(self, record):
        self._append_row(self.CONTENT_SHEET, record.to_row())

    def replace_content(self, records):
        self._write_all(self.CONTENT_SHEET, ContentRecord.HEADER, [r.to_row() for r in records])

    def append_history(self, record):
        self._append_row(self.HISTORY_SHEET, record.to_row())

    def list_history(self):
        return [ResearchHistoryRecord.from_row(r) for r in self._read_rows(self.HISTORY_SHEET)]
